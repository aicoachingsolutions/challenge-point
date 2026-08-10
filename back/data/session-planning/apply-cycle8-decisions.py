"""
Apply Christian's Cycle 8 Canonical Decisions to the canonical Session Planning workbook.

This is the FIRST script here that writes the canonical workbook rather than a review copy.
Christian's Cycle 8 email put workbook implementation and validation on our side: "Rather than
editing the workbooks ourselves, we're leaving workbook implementation and validation in your
hands." So these are his decisions, applied — not proposals.

Run from the repo root:
    python back/data/session-planning/apply-cycle8-decisions.py
    python back/data/session-planning/project-workbook.py

WHAT IS APPLIED

  §4 Engine Translation — approved as the Primary Representative Game Problem Translation for
     Pilot RC1. The nine mappings we proposed become canonical.

  §5 Intentional Runtime Gaps — A01 and A04 stay unresolved. His words: "No placeholder mappings
     should be introduced." They are written as EMPTY, not "TBD", because TBD reads as "not looked
     at yet" and these were looked at and deliberately left.

  §2 Entry Language — the 46 verified additions are merged.

  §3 Judgment Phrases — resolved by his stated coaching intent, not by parser behaviour. D02 owns
     collective ball-winning; D03 owns explicit individual defending. His four D03 phrases are
     added even though some do not yet route, because Entry Language is navigation rather than
     routing and the phrases are his authored coach language.

WHAT IS DELIBERATELY NOT APPLIED

  LG-012 "Breaking Lines". The Canonical Decisions document §1 says to add it. His covering email
  and the Cycle 8 README both say the opposite — "No additional Learning Goals are introduced
  during this cycle", and the concept is "better represented through Entry Language, translation,
  and future knowledge population than through a new planning object". Two sources say no, one says
  yes, and the README's own deliverables table drops LG-012 from the list the Canonical Decisions
  table still carries as unsent.

  Independently of which is current: the LG-012 artifact was never sent (Canonical Decisions
  deliverables table marks it "Sent to Joe" unchecked). Adding it would mean inventing its Phase,
  Display Order, Coach Definition, "Choose This When", Common Misclassification and Practice
  Situations — authoring his knowledge, which is exactly what the Governance Standard forbids.
  So it is raised with him rather than guessed at.

  "defend" and "defending" as judgment phrases. His §3 distinction separates collective ball-winning
  (D02) from EXPLICIT individual defending (D03). Bare "defend" is neither: it is a general
  defensive word that could as easily mean Stay Organized or Delay the Attack. Assigning it to D02
  would assert that "defend" means ball-winning, which is a coaching judgement he did not make.
"""

import json
import pathlib

import openpyxl

HERE = pathlib.Path(__file__).parent
WORKBOOK = HERE / "session-planning-model.rc1.xlsx"
TRANSLATION = HERE / "engine-translation.proposed.json"
ENTRY = HERE / "entry-language.proposed.json"

HEADER_ROW = 1

# §3 — resolved by his stated coaching intent. The six he named explicitly, plus the two
# "win the ball" variants that carry the same collective ball-winning intent as "win it back".
JUDGMENT_ASSIGNMENTS = {
    "press": "D02",
    "pressing": "D02",
    "counterpress": "D02",
    "regain": "D02",
    "win it back": "D02",
    "turnover": "D02",
    "win the ball": "D02",
    "winning the ball": "D02",
}

# §3 — phrases Christian authored for D03 that our extraction did not contain.
AUTHORED_D03_PHRASES = ["isolate attacker", "stop your player", "defend the dribbler"]

# Left unassigned on purpose — see the module docstring.
DEFERRED_JUDGMENT = ["defend", "defending"]


def apply_engine_translation(worksheet):
    with TRANSLATION.open(encoding="utf-8") as handle:
        rows = {r["learningGoalId"]: r for r in json.load(handle)}

    headers = [c.value for c in worksheet[HEADER_ROW]]
    index = {h: i + 1 for i, h in enumerate(headers) if h}

    applied, unresolved = 0, []
    for r in range(HEADER_ROW + 1, worksheet.max_row + 1):
        goal_id = worksheet.cell(row=r, column=index["Learning Goal ID"]).value
        if goal_id is None:
            continue
        row = rows.get(str(goal_id).strip())
        if row is None:
            continue

        primary = worksheet.cell(row=r, column=index["Primary GP IDs"])
        secondary = worksheet.cell(row=r, column=index["Secondary GP IDs"])

        if row["primary"]:
            primary.value = "; ".join(row["primary"])
            # §4: secondary Game Problems are explicitly a FUTURE capability, so RC1 records the
            # primary only. Writing a secondary now would claim a relationship he has not approved.
            secondary.value = None
            worksheet.cell(row=r, column=index["Notes"], value="Approved RC1 — Primary Representative Game Problem Translation (Cycle 8).")
            applied += 1
        else:
            primary.value = None
            secondary.value = None
            worksheet.cell(
                row=r,
                column=index["Notes"],
                value="Intentional runtime gap (Cycle 8 §5) — no placeholder mapping. Runtime realization opportunity, not a knowledge deficiency.",
            )
            unresolved.append(str(goal_id))

    return applied, unresolved


def apply_entry_language(worksheet):
    with ENTRY.open(encoding="utf-8") as handle:
        draft = json.load(handle)

    existing = set()
    last_row = HEADER_ROW
    for r in range(HEADER_ROW + 1, worksheet.max_row + 1):
        phrase = worksheet.cell(row=r, column=1).value
        if phrase is None or not str(phrase).strip():
            continue
        existing.add(str(phrase).strip().lower())
        last_row = r

    additions = []
    for row in draft["proposed"]:
        additions.append((row["phrase"], row["learningGoalId"]))
    for phrase, goal_id in JUDGMENT_ASSIGNMENTS.items():
        additions.append((phrase, goal_id))
    for phrase in AUTHORED_D03_PHRASES:
        additions.append((phrase, "D03"))

    written = 0
    for phrase, goal_id in additions:
        key = phrase.strip().lower()
        if key in existing:
            continue
        existing.add(key)
        last_row += 1
        worksheet.cell(row=last_row, column=1, value=phrase)
        worksheet.cell(row=last_row, column=2, value=goal_id)
        written += 1

    return written


def main():
    workbook = openpyxl.load_workbook(WORKBOOK)

    applied, unresolved = apply_engine_translation(workbook["Engine Translation"])
    print(f"  Engine Translation: {applied} approved mappings applied")
    print(f"    intentional gaps left empty: {', '.join(unresolved)}")

    written = apply_entry_language(workbook["Entry Language"])
    print(f"  Entry Language: {written} phrases merged")
    print(f"    judgment phrases resolved by coaching intent: {len(JUDGMENT_ASSIGNMENTS)}")
    print(f"    authored D03 phrases added: {len(AUTHORED_D03_PHRASES)}")
    print(f"    deferred, outside his stated distinction: {', '.join(DEFERRED_JUDGMENT)}")

    workbook.save(WORKBOOK)
    print(f"Saved {WORKBOOK.name} — CANONICAL update, not a review copy.")
    print("  LG-012 Breaking Lines NOT added — sources conflict and the artifact was not sent. Raised with Christian.")


if __name__ == "__main__":
    main()
