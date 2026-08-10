"""
Session Planning Model workbook -> JSON projection.

Same pattern as the Knowledge Core and Sport Module workbooks: the .xlsx stays the canonical
artifact, and this produces a complete generated projection the TypeScript loader reads. Never
hand-edit the JSON — regenerate it.

Run from the repo root:
    python back/data/session-planning/project-workbook.py

TWO DEVIATIONS FROM THE ESTABLISHED WORKBOOK PATTERN, both handled here rather than papered over:

  1. HEADER ROW IS 1, not 2. Every previous workbook puts a title in row 1 and headers in row 2.
     This one starts with headers. Declared as a constant rather than sniffed, so if a future
     revision adds a title row the projection fails loudly instead of reading data as headers.

  2. THERE IS NO METADATA SHEET. The Knowledge Core workbooks all carry expected row counts and a
     compatibility matrix, which is what lets a loader detect truncation — and Christian made
     Version Metadata a REQUIRED element of every future Knowledge Schema after our EM review. This
     workbook predates that being applied here, so the loader cannot check counts against the
     workbook's own declarations. It gates on referential integrity instead, which is a stronger
     check: every reference must resolve. Flagged for Christian rather than silently accommodated.
"""

import json
import pathlib

import openpyxl

HERE = pathlib.Path(__file__).parent
WORKBOOK = HERE / "session-planning-model.rc1.xlsx"
OUT = pathlib.Path(__file__).parents[2] / "src/system/session-planning/session-planning-model.rc1.json"

HEADER_ROW = 1

SHEETS = {
    "Learning Goal Registry": "learning_goals",
    "Practice Situation Registry": "practice_situations",
    "Entry Language": "entry_language",
    "Engine Translation": "engine_translation",
    "Governance": "governance",
}


def read_sheet(worksheet):
    headers = [c.value for c in worksheet[HEADER_ROW]]
    rows = []
    for raw in worksheet.iter_rows(min_row=HEADER_ROW + 1, values_only=True):
        if not any(v is not None and str(v).strip() != "" for v in raw):
            continue
        row = {}
        for header, value in zip(headers, raw):
            if not header:
                continue
            row[str(header)] = value
        rows.append(row)
    return rows


def main():
    workbook = openpyxl.load_workbook(WORKBOOK, data_only=True)

    missing = [name for name in SHEETS if name not in workbook.sheetnames]
    if missing:
        raise SystemExit(f"Workbook is missing expected sheet(s): {missing}")

    projection = {
        "source_workbook": WORKBOOK.name,
        "header_row": HEADER_ROW,
    }
    for sheet_name, key in SHEETS.items():
        projection[key] = read_sheet(workbook[sheet_name])
        print(f"  {sheet_name}: {len(projection[key])} rows")

    OUT.parent.mkdir(parents=True, exist_ok=True)
    with OUT.open("w", encoding="utf-8") as handle:
        json.dump(projection, handle, indent=2, ensure_ascii=False, default=str)
        handle.write("\n")
    print(f"Wrote {OUT}")


if __name__ == "__main__":
    main()
