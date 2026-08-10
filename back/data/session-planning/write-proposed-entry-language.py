"""
Write proposed Entry Language into a COPY of Christian's workbook, for review.

Run after the TS proposer:
    npx ts-node --files -r tsconfig-paths/register ./src/scripts/propose-entry-language.ts
    python back/data/session-planning/write-proposed-entry-language.py

THREE SHEETS, BECAUSE THERE ARE THREE DIFFERENT ASKS AND MIXING THEM WOULD WASTE HIS TIME:

  Entry Language          — his existing rows plus the phrases that map to exactly one Learning
                            Goal. Same two columns as his schema, so it can be accepted wholesale.
  PROPOSED - Judgement    — phrases whose signal group is shared by several Learning Goals. Which
                            one they mean is a coaching call, and guessing would put an invented
                            judgement into his canonical model.
  PROPOSED - Uncovered    — the headline. Vocabulary the engine supports that NO Learning Goal
                            reaches, grouped into concept areas. These are not phrases to file;
                            they are candidate Learning Goals that do not exist yet.

His canonical schema is never widened — proposals live on new sheets, and the Entry Language sheet
keeps exactly the columns it already had.
"""

import collections
import json
import pathlib
import shutil

import openpyxl

HERE = pathlib.Path(__file__).parent
SOURCE = HERE / "session-planning-model.rc1.xlsx"
PROPOSAL = HERE / "session-planning-model.rc1-PROPOSED-entry-language.xlsx"
DRAFT = HERE / "entry-language.proposed.json"

HEADER_ROW = 1

# Signal group -> the coaching concept it represents, for the uncovered-areas sheet. These are
# descriptions of what the ENGINE already supports, not proposed Learning Goal names — naming them
# would be authoring his model for him.
CONCEPT_AREAS = {
    "I_defensive": "General defensive intent (protect, deny, contain, compact, screen)",
    "I_defensive_press": "Active ball-winning pressure",
    "C_spacing_support": "Spacing, support angles, width and switching play",
    "K_information": "Perception and decision-making (reading, scanning, disguise, blind side)",
    "D_break_lines": "Breaking lines and penetrating forward",
    "A_touch_receiving": "First touch and receiving under pressure",
    "G_overload": "Numerical advantage and overloads",
}


def main():
    if not DRAFT.exists():
        raise SystemExit(f"{DRAFT.name} not found — run propose-entry-language.ts first.")

    with DRAFT.open(encoding="utf-8") as handle:
        draft = json.load(handle)

    shutil.copyfile(SOURCE, PROPOSAL)
    workbook = openpyxl.load_workbook(PROPOSAL)

    # 1. Append the unambiguous rows to the existing sheet, preserving its two columns exactly.
    entry = workbook["Entry Language"]
    next_row = entry.max_row + 1
    for row in draft["proposed"]:
        entry.cell(row=next_row, column=1, value=row["phrase"])
        entry.cell(row=next_row, column=2, value=row["learningGoalId"])
        next_row += 1

    # 2. Phrases needing a coaching call.
    judgement = workbook.create_sheet("PROPOSED - Judgement")
    judgement.append(["Coach Phrase", "Candidate Learning Goal IDs", "Why this needs your call"])
    for row in draft["review"]:
        if row["candidates"]:
            judgement.append([row["phrase"], row["candidates"], row["reason"]])

    # 3. THE HEADLINE — supported vocabulary with no home in the planning model.
    uncovered = workbook.create_sheet("PROPOSED - Uncovered")
    uncovered.append(["Concept area the engine supports", "Phrases", "Example coach phrases"])
    grouped = collections.defaultdict(list)
    for row in draft["review"]:
        if row["candidates"]:
            continue
        # The reason carries the signal group; recover it rather than re-deriving.
        group = row["reason"].split("Signal group ")[1].split(" ")[0]
        grouped[group].append(row["phrase"])

    for group, phrases in sorted(grouped.items(), key=lambda kv: -len(kv[1])):
        uncovered.append(
            [
                CONCEPT_AREAS.get(group, group),
                len(phrases),
                ", ".join(phrases[:12]) + (" …" if len(phrases) > 12 else ""),
            ]
        )

    workbook.save(PROPOSAL)
    print(f"Wrote {PROPOSAL.name}")
    print(f"  Entry Language: +{len(draft['proposed'])} rows appended to his existing {next_row - 1 - len(draft['proposed']) - HEADER_ROW}")
    print(f"  PROPOSED - Judgement: {sum(1 for r in draft['review'] if r['candidates'])} phrases")
    print(f"  PROPOSED - Uncovered: {len(grouped)} concept areas, {sum(len(v) for v in grouped.values())} phrases")
    print("  Source workbook untouched — review copy only.")


if __name__ == "__main__":
    main()
