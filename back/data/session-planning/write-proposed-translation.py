"""
Write the proposed Engine Translation into a COPY of Christian's workbook, for his review.

Run after the TS proposer:
    npx ts-node --files -r tsconfig-paths/register ./src/scripts/propose-engine-translation.ts
    python back/data/session-planning/write-proposed-translation.py

WRITES TO A SEPARATE FILE, NEVER THE INGESTED WORKBOOK. The ingested copy is the canonical artifact
we load from; a proposal is not canonical until Christian accepts it. Keeping them as different
files means there is no moment where a draft could be mistaken for his decision — which matters more
here than convenience, because the mapping is a coaching judgement that belongs to him under his own
Governance Standard.

Only the Engine Translation sheet is touched. Every other sheet is copied through untouched.
"""

import json
import pathlib
import shutil

import openpyxl

HERE = pathlib.Path(__file__).parent
SOURCE = HERE / "session-planning-model.rc1.xlsx"
PROPOSAL = HERE / "session-planning-model.rc1-PROPOSED-engine-translation.xlsx"
DRAFT = HERE / "engine-translation.proposed.json"

HEADER_ROW = 1
SHEET = "Engine Translation"


def main():
    if not DRAFT.exists():
        raise SystemExit(f"{DRAFT.name} not found — run propose-engine-translation.ts first.")

    with DRAFT.open(encoding="utf-8") as handle:
        rows = {row["learningGoalId"]: row for row in json.load(handle)}

    shutil.copyfile(SOURCE, PROPOSAL)
    workbook = openpyxl.load_workbook(PROPOSAL)
    worksheet = workbook[SHEET]

    headers = [c.value for c in worksheet[HEADER_ROW]]
    index = {h: i + 1 for i, h in enumerate(headers) if h}

    written = 0
    gaps = 0
    for r in range(HEADER_ROW + 1, worksheet.max_row + 1):
        goal_id = worksheet.cell(row=r, column=index["Learning Goal ID"]).value
        if goal_id is None:
            continue
        row = rows.get(str(goal_id).strip())
        if row is None:
            continue

        # NOTE: assign through .value. openpyxl's cell(..., value=None) is a NO-OP, so the sheet's
        # existing "TBD" placeholders survive a "clear" that looks like it worked.
        primary_cell = worksheet.cell(row=r, column=index["Primary GP IDs"])
        secondary_cell = worksheet.cell(row=r, column=index["Secondary GP IDs"])

        if row["primary"]:
            primary_cell.value = "; ".join(row["primary"])
            secondary_cell.value = "; ".join(row["secondary"]) if row["secondary"] else None
            written += 1
        else:
            # Left blank on purpose. A plausible-looking ID here would bury the finding.
            primary_cell.value = None
            secondary_cell.value = None
            gaps += 1

        worksheet.cell(row=r, column=index["Notes"], value=row["notes"])

    workbook.save(PROPOSAL)
    print(f"Wrote {PROPOSAL.name}")
    print(f"  {written} rows proposed, {gaps} left blank as gaps")
    print("  Source workbook untouched — this is a review copy, not a canonical update.")


if __name__ == "__main__":
    main()
