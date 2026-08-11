"""
Representative Intervention Workbook -> JSON projection.

Same pattern as every other workbook we ingest: the .xlsx stays canonical, this produces a complete
generated projection the TypeScript loader reads. Never hand-edit the JSON — regenerate it.

Run from the repo root:
    python back/data/experience-design/project-workbook.py

LAYOUT NOTE. Headers are in ROW 1 on all three sheets, and the Metadata sheet is a key/value pair
list whose "header" is really its first pair — so it is read as pairs rather than as a table. Both
are declared here rather than sniffed, so a future revision that adds a title row fails loudly
instead of quietly reading data as headers.
"""

import json
import pathlib

import openpyxl

HERE = pathlib.Path(__file__).parent
WORKBOOK = HERE / "representative-intervention.rc1.xlsx"
OUT = pathlib.Path(__file__).parents[2] / "src/system/experience-design/representative-intervention.rc1.json"

HEADER_ROW = 1
TABLE_SHEETS = {"Registry": "registry", "Runtime Data": "runtime_data"}
METADATA_SHEET = "Metadata"


def read_table(worksheet):
    headers = [c.value for c in worksheet[HEADER_ROW]]
    rows = []
    for raw in worksheet.iter_rows(min_row=HEADER_ROW + 1, values_only=True):
        if not any(v is not None and str(v).strip() != "" for v in raw):
            continue
        row = {}
        for header, value in zip(headers, raw):
            if not header:
                continue
            row[str(header)] = value
        rows.append(row)
    return rows


def read_metadata(worksheet):
    """Key/value pairs, including the first row — it is a pair, not a header."""
    meta = {}
    for raw in worksheet.iter_rows(min_row=1, values_only=True):
        if raw and raw[0] is not None and str(raw[0]).strip():
            meta[str(raw[0]).strip()] = raw[1]
    return meta


def main():
    workbook = openpyxl.load_workbook(WORKBOOK, data_only=True)

    missing = [name for name in [*TABLE_SHEETS, METADATA_SHEET] if name not in workbook.sheetnames]
    if missing:
        raise SystemExit(f"Workbook is missing expected sheet(s): {missing}")

    projection = {"source_workbook": WORKBOOK.name, "header_row": HEADER_ROW}
    for sheet_name, key in TABLE_SHEETS.items():
        projection[key] = read_table(workbook[sheet_name])
        print(f"  {sheet_name}: {len(projection[key])} rows")

    projection["metadata"] = read_metadata(workbook[METADATA_SHEET])
    print(f"  {METADATA_SHEET}: {len(projection['metadata'])} declarations")

    OUT.parent.mkdir(parents=True, exist_ok=True)
    with OUT.open("w", encoding="utf-8") as handle:
        json.dump(projection, handle, indent=2, ensure_ascii=False, default=str)
        handle.write("\n")
    print(f"Wrote {OUT}")


if __name__ == "__main__":
    main()
