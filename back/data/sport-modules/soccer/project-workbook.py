"""
Soccer Module workbook -> JSON projection.

Same pattern proven six times on the canonical Knowledge Core workbooks: the .xlsx stays the
canonical artifact, and this produces a complete generated projection the TypeScript loader reads.
Never hand-edit the JSON — regenerate it.

Run from the repo root:
    python back/data/sport-modules/soccer/project-workbook.py

Reads the per-sheet header rows from the workbook's own Metadata sheet rather than assuming, so a
future workbook that moves a header does not silently produce empty rows.
"""

import json
import pathlib

import openpyxl

HERE = pathlib.Path(__file__).parent
WORKBOOK = HERE / "soccer-module.rc1-v3.xlsx"
OUT = pathlib.Path(__file__).parents[3] / "src/system/sport-module/soccer-module.rc1-v3.json"

DATA_SHEETS = ["Vocabulary", "Lenses", "Game Forms", "Realizations", "Coverage"]


def read_metadata(worksheet):
    meta = {}
    for r in range(3, worksheet.max_row + 1):
        key = worksheet.cell(row=r, column=1).value
        if key is None:
            continue
        meta[str(key)] = worksheet.cell(row=r, column=2).value
    return meta


def sheet_key(sheet_name):
    return sheet_name.lower().replace(" ", "_")


def read_sheet(worksheet, header_row):
    headers = [c.value for c in worksheet[header_row]]
    rows = []
    for raw in worksheet.iter_rows(min_row=header_row + 1, values_only=True):
        if not any(v is not None and str(v).strip() != "" for v in raw):
            continue
        row = {}
        for header, value in zip(headers, raw):
            if not header:
                continue
            row[str(header)] = value
        rows.append(row)
    return rows


def main():
    workbook = openpyxl.load_workbook(WORKBOOK, data_only=True)
    metadata = read_metadata(workbook["Metadata"])

    projection = {"metadata": metadata}
    for sheet_name in DATA_SHEETS:
        key = sheet_key(sheet_name)
        header_row = int(metadata.get(f"{key}_header_row", 2))
        projection[key] = read_sheet(workbook[sheet_name], header_row)
        print(f"  {sheet_name}: {len(projection[key])} rows (header row {header_row})")

    OUT.parent.mkdir(parents=True, exist_ok=True)
    with OUT.open("w", encoding="utf-8") as handle:
        json.dump(projection, handle, indent=2, ensure_ascii=False, default=str)
        handle.write("\n")
    print(f"Wrote {OUT}")


if __name__ == "__main__":
    main()
