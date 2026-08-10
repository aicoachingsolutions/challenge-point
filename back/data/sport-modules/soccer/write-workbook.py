"""
Soccer Module workbook writer — one-time migration step.

Reads the JSON emitted by the TypeScript extractors and writes it into the governed workbook,
updating the Metadata expected-row counts so the loader's integrity gate has something to check.

Run from the repo root, after the TS extractor:
    python back/data/sport-modules/soccer/write-workbook.py

Deliberately conservative:
  * writes only sheets it has data for, so a partial extraction never blanks a populated sheet;
  * refuses to write a column the workbook does not declare, rather than silently adding one — an
    unrecognised column means the extractor and the schema have diverged, which is exactly the
    failure this whole exercise exists to prevent;
  * updates <sheet>_expected_rows so the loader can detect truncation later.
"""

import json
import pathlib
import sys

import openpyxl

HERE = pathlib.Path(__file__).parent
WORKBOOK = HERE / "soccer-module.rc1-v3.xlsx"
HEADER_ROW = 2  # declared by the workbook's own Metadata sheet

# sheet name -> extracted JSON file
# RETIRED SHEETS. Game Forms and Realizations are now AUTHORED IN THE WORKBOOK — Christian has
# populated coach vocabulary, canonical archetype identifiers and Game Problem mappings that exist
# nowhere in the code. Re-running their extractors would silently destroy that work, so they are
# deliberately not listed here. The extraction scripts remain as the audit trail for how the rows
# were first derived; they must not be used to rewrite these sheets again.
#
# Lenses is still generated, because Christian has not authored into it.
SOURCES = {
    "Lenses": HERE / "lenses.extracted.json",
    # Vocabulary: RETIRED 2026-08-08, immediately after its first and only generation.
    #
    # Retired BEFORE Christian authors into it rather than after, which is the whole point. The sheet
    # exists so vocabulary becomes an edit instead of a deploy; the moment he adds a phrase, a rerun
    # of this writer would silently destroy it. Game Forms and Realizations taught us that, and there
    # is no reason to leave the window open a second time. `extract-soccer-vocabulary.ts` remains as
    # the audit trail for how the 175 rows were derived — it must not rewrite this sheet again.
    # Realization Banks is generated for the same reason Lenses is: the entries were derived from
    # code and Christian has not authored into the sheet yet. When he does, retire it as above.
    "Realization Banks": HERE / "realization-banks.extracted.json",
    # Coverage is DERIVED from the live engine, so it is regenerated rather than authored — a
    # hand-edited coverage claim would silently stop matching what the engine actually does.
    "Coverage": HERE / "coverage.extracted.json",
}


def load_rows(path: pathlib.Path):
    if not path.exists():
        return None
    with path.open(encoding="utf-8") as handle:
        return json.load(handle)


def write_sheet(worksheet, rows):
    headers = [c.value for c in worksheet[HEADER_ROW]]
    index = {h: i for i, h in enumerate(headers) if h}

    unknown = sorted({k for row in rows for k in row} - set(index))
    if unknown:
        raise SystemExit(
            f"[{worksheet.title}] extractor produced columns the workbook does not declare: {unknown}\n"
            "The extractor and the schema have diverged — fix one of them rather than writing anyway."
        )

    for offset, row in enumerate(rows):
        target = HEADER_ROW + 1 + offset
        for key, value in row.items():
            worksheet.cell(row=target, column=index[key] + 1, value=value if value != "" else None)

    # CLEAR THE TAIL. A sheet that SHRINKS leaves orphan rows behind, and they are invisible: the
    # projection reads every non-empty row, so the loader sees more rows than were written and the
    # integrity gate — which compares against the count we just wrote — disagrees for a reason that
    # points nowhere. Caught 2026-08-08 when a parser fix took vocabulary from 175 rows to 173 and
    # two stale rows survived, still asserting routing the parser no longer does.
    # delete_rows, NOT blanking the cells. Setting values to None leaves the cells in the sheet and
    # the projection still counts the row, so the orphans survive a "clear" that looks like it worked
    # — which is how this defect hid in the first place.
    first_orphan = HEADER_ROW + 1 + len(rows)
    populated = [
        r
        for r in range(first_orphan, worksheet.max_row + 1)
        if any(worksheet.cell(row=r, column=c).value not in (None, "") for c in range(1, len(headers) + 1))
    ]
    if populated:
        worksheet.delete_rows(first_orphan, max(populated) - first_orphan + 1)
        print(f"    cleared {len(populated)} orphan row(s) left by a shrinking sheet")

    return len(rows)


def update_metadata(worksheet, counts):
    keys = {}
    for r in range(3, worksheet.max_row + 1):
        key = worksheet.cell(row=r, column=1).value
        if key:
            keys[str(key)] = r

    for sheet_name, count in counts.items():
        metadata_key = f"{sheet_name.lower().replace(' ', '_')}_expected_rows"
        if metadata_key in keys:
            worksheet.cell(row=keys[metadata_key], column=2, value=count)
        else:
            print(f"  ! no metadata key '{metadata_key}' to update", file=sys.stderr)


def main():
    workbook = openpyxl.load_workbook(WORKBOOK)
    counts = {}

    for sheet_name, source in SOURCES.items():
        rows = load_rows(source)
        if rows is None:
            print(f"  - {sheet_name}: no extract yet, leaving as-is")
            continue
        counts[sheet_name] = write_sheet(workbook[sheet_name], rows)
        print(f"  + {sheet_name}: wrote {counts[sheet_name]} rows")

    if counts:
        update_metadata(workbook["Metadata"], counts)
        workbook.save(WORKBOOK)
        print(f"Saved {WORKBOOK.name}")
    else:
        print("Nothing to write.")


if __name__ == "__main__":
    main()
