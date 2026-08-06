"""
Soccer Module workbook — ONE-TIME schema change adding the Realization Banks resource.

Christian approved modelling the realization banks as a normalized one-to-many resource rather than
flattening them into a workbook column (2026-08-05). This performs the structural half of that:

  * creates the "Realization Banks" sheet with its declared header row, matching the layout of every
    other sheet (title in row 1, headers in row 2, data from row 3);
  * adds the three Metadata declarations the loader's integrity gate reads
    (realization_banks_sheet_name / _header_row / _expected_rows);
  * populates `realization_bank_id` on the four owning Realizations rows, completing an FK the sheet
    already declared but had nothing to point at.

The bank ROWS are written by the normal `write-workbook.py` path, not here — this script only makes
the shape exist. It is idempotent: rerunning it does not duplicate the sheet, the metadata keys, or
the FKs.

Run from the repo root, once:
    python back/data/sport-modules/soccer/add-realization-banks-sheet.py
"""

import json
import pathlib

import openpyxl

HERE = pathlib.Path(__file__).parent
WORKBOOK = HERE / "soccer-module.rc1-v3.xlsx"
OWNERS = HERE / "realization-bank-owners.extracted.json"

SHEET_NAME = "Realization Banks"
TITLE = "Soccer Module — Realization Banks Resource (RC1 Candidate)"
HEADER_ROW = 2

# Ordinal sits third, right beside the keys it orders — a reader scanning the sheet should not have
# to hunt for the column that decides which entry gets designated.
HEADERS = [
    "realization_bank_entry_id",
    "realization_bank_id",
    "realization_id",
    "bank_ordinal",
    "realization_text",
    "universal_concept_type",
    "status",
    "introduced_version",
    "last_verified_version",
    "provenance",
    "notes",
]

# Inserted after the Realizations block so the Metadata sheet keeps its sheet-by-sheet reading order.
METADATA_AFTER = "realizations_expected_rows"
METADATA_ROWS = [
    ("realization_banks_sheet_name", SHEET_NAME),
    ("realization_banks_header_row", HEADER_ROW),
    ("realization_banks_expected_rows", 0),
]


def ensure_sheet(workbook):
    if SHEET_NAME in workbook.sheetnames:
        print(f"  = sheet '{SHEET_NAME}' already present")
        return workbook[SHEET_NAME]

    # Placed before Metadata so the data sheets stay contiguous.
    index = workbook.sheetnames.index("Metadata")
    worksheet = workbook.create_sheet(SHEET_NAME, index)
    worksheet.cell(row=1, column=1, value=TITLE)
    for column, header in enumerate(HEADERS, start=1):
        worksheet.cell(row=HEADER_ROW, column=column, value=header)
    print(f"  + created sheet '{SHEET_NAME}' with {len(HEADERS)} columns")
    return worksheet


def ensure_metadata(worksheet):
    existing = {}
    for r in range(3, worksheet.max_row + 1):
        key = worksheet.cell(row=r, column=1).value
        if key:
            existing[str(key)] = r

    missing = [(k, v) for k, v in METADATA_ROWS if k not in existing]
    if not missing:
        print("  = metadata declarations already present")
        return

    anchor = existing.get(METADATA_AFTER)
    if anchor is None:
        raise SystemExit(f"Metadata has no '{METADATA_AFTER}' row to insert after; workbook shape has changed.")

    worksheet.insert_rows(anchor + 1, amount=len(missing))
    for offset, (key, value) in enumerate(missing):
        worksheet.cell(row=anchor + 1 + offset, column=1, value=key)
        worksheet.cell(row=anchor + 1 + offset, column=2, value=value)
    print(f"  + added {len(missing)} metadata declarations after '{METADATA_AFTER}'")


def ensure_foreign_keys(worksheet):
    if not OWNERS.exists():
        raise SystemExit(f"{OWNERS.name} not found — run extract-soccer-realization-banks.ts first.")
    with OWNERS.open(encoding="utf-8") as handle:
        owners = {o["realization_id"]: o["realization_bank_id"] for o in json.load(handle)}

    headers = [c.value for c in worksheet[HEADER_ROW]]
    index = {h: i + 1 for i, h in enumerate(headers) if h}
    id_column = index["realization_id"]
    fk_column = index["realization_bank_id"]

    written = 0
    seen = set()
    for r in range(HEADER_ROW + 1, worksheet.max_row + 1):
        realization_id = worksheet.cell(row=r, column=id_column).value
        if realization_id is None:
            continue
        bank_id = owners.get(str(realization_id))
        if bank_id is None:
            continue
        seen.add(str(realization_id))
        cell = worksheet.cell(row=r, column=fk_column)
        if cell.value == bank_id:
            continue
        cell.value = bank_id
        written += 1

    # A bank whose owner is not on the sheet is a dangling reference — say so rather than write a
    # resource nothing can reach. Checked on rows SEEN, not on rows written, so an already-populated
    # workbook still catches a missing owner.
    unmatched = sorted(set(owners) - seen)
    if unmatched:
        raise SystemExit(f"Owning realization rows not found on the sheet: {unmatched}")
    print(f"  {'+' if written else '='} realization_bank_id set on {written} of {len(seen)} owning rows")


def main():
    workbook = openpyxl.load_workbook(WORKBOOK)
    ensure_sheet(workbook)
    ensure_metadata(workbook["Metadata"])
    ensure_foreign_keys(workbook["Realizations"])
    workbook.save(WORKBOOK)
    print(f"Saved {WORKBOOK.name}")


if __name__ == "__main__":
    main()
